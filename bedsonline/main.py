import operator
from time import sleep
from telnetlib import EC
from bs4 import BeautifulSoup
from openpyxl.styles import PatternFill, Font
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.common.exceptions import NoSuchElementException, TimeoutException
# todo: EXPECTED CONDITIONS https://selenium-python.readthedocs.io/waits.html
# https://stackoverflow.com/questions/58743549/attributeerror-bytes-object-has-no-attribute-element-to-be-clickable
# The import from telnetlib import EC. You need to import expected_conditions and use it as EC
# from selenium.webdriver.support import expected_conditions as EC...
from selenium.webdriver.support import expected_conditions as EC
from bedsonline import targets as t
from openpyxl import load_workbook
from openpyxl import Workbook
import requests
import os

global mode


class App:
    def __init__(self, username='BUSIN95C', password='020906Sm', target_continent='2', target_country_col='3',
                 target_country_row='1', target_city_col='4', target_city_row='94', path='//home/jmartorell/Booking'):
        self.username = username
        self.password = password
        self.target_continent = target_continent
        self.target_country_col = target_country_col
        self.target_country_row = target_country_row
        self.target_city_col = target_city_col
        self.target_city_row = target_city_row
        self.all_positions = []
        self.all_hotels = []
        self.all_prices = []
        self.all_zones = []
        self.display = []
        self.cheap = []
        self.index = ""
        self.data = {}
        self.path = path
        self.driver = webdriver.Firefox(
            executable_path="/usr/local/bin/geckodriver")  # Change this to your FirefoxDriver path.
        self.error = False
        self.main_url = 'https://www.bedsonline.com/home/es-es'
        self.all_images = []
        self.driver.get(self.main_url)
        sleep(1)
        self.log_in()
        if self.error is False:
            sleep(2)  # fixme: implicit wait!
            self.search_target_profile()
        if self.error is False:
            self.scroll_down()
        if self.error is False:
            if not os.path.exists(path):
                os.mkdir(path)
            self.file_manager()
        # sleep(10)
        self.driver.close()

    def log_in(self, ):
        try:
            # todo: dealing with popup windows
            # cookies popup
            print('\nClosing cookies window ...')
            WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((
                By.XPATH,
                '//div/div/div[2]/button[1]'))).click()
            sleep(1)

        except Exception:
            self.error = True
            print('Unable to close popup window')
        else:
            try:
                # Press button to access the login fields
                sleep(1)
                login_button = self.driver.find_element_by_xpath(
                    '//section[2]/div[1]/div/section[2]/a[1]')
                login_button.click()
                sleep(1)

                # todo: switch window
                #  https://sqa.stackexchange.com/questions/13792/how-to-proceed-after-clicking-a-link-to-new-page-in-selenium-in-python
                window_after = self.driver.window_handles[1]
                self.driver.switch_to.window(window_after)
                print('Logging in with username and password ...')

                WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((
                    By.XPATH, '//*[@id="username"]')))
                user_name_input = self.driver.find_element_by_xpath('//*[@id="username"]')
                user_name_input.send_keys(self.username)
                sleep(1)

                WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((
                    By.XPATH, '//*[@id="password"]')))
                password_input = self.driver.find_element_by_xpath('//*[@id="password"]')
                password_input.send_keys(self.password)
                sleep(1)

                # user_name_input.submit()
                self.driver.find_element_by_xpath('//form/div[3]/button').submit()
                sleep(1)

            except Exception as e:
                print(e)
                self.error = True

    def search_target_profile(self):
        try:
            # todo: filling the search bar
            # search_bar.send_keys(self.target_username).click # this line is wrong, read below please
            # fixme: you cannot enter text directly use the autocomplete square icon to the right of the field
            WebDriverWait(self.driver, 100).until(EC.visibility_of_element_located((
                By.XPATH, '//section/div/form/fieldset[1]/div/a/span'))).click()

            # todo: https://dzone.com/articles/perform-actions-using-javascript-in-python-seleniu
            #  Method 1: Executing JavaScript at Document Root Level
            #  javaScript = "document.getElementsByClassName('ui-dialog-titlebar-close ui-corner-all')[0].click();"
            #  self.driver.execute_script(javaScript)
            #  Method 2: Executing JavaScript at Element Level:
            print("Manipulating search engine ...")
            try:
                wait = WebDriverWait(self.driver, 100)
                wait.until(
                    EC.element_to_be_clickable(
                        (By.CSS_SELECTOR, "#continent-picker-tab > li:nth-child"
                                          "(" + self.target_continent + ") > a:nth-child(1)")))
            except TimeoutException:
                print("Searching took too much time!")

            if self.target_continent != '4':
                picker = WebDriverWait(self.driver, 100).until(EC.visibility_of_element_located((
                    By.CSS_SELECTOR,
                    "#continent-picker-tab > li:nth-child(" + self.target_continent + ") > a:nth-child(1)")))
                self.driver.execute_script("arguments[0].click();", picker)
                sleep(1)
            picker = WebDriverWait(self.driver, 100).until(EC.visibility_of_element_located((
                By.XPATH,
                "//section/article[1]/div/ul[" + self.target_country_col + "]/li[" + self.target_country_row + "]/a")))
            self.driver.execute_script("arguments[0].click();", picker)
            sleep(1)
            picker = WebDriverWait(self.driver, 100).until(EC.visibility_of_element_located((
                By.XPATH,
                "//section/article[2]/div/ul[" + self.target_city_col + "]/li[" + self.target_city_row + "]/a")))
            self.driver.execute_script("arguments[0].click();", picker)
            sleep(1)
            picker = WebDriverWait(self.driver, 100).until(EC.visibility_of_element_located((
                By.XPATH,
                "//article[3]/div/ul[1]/li[1]/a")))
            self.driver.execute_script("arguments[0].click();", picker)
            sleep(1)

            # search button
            login_button = WebDriverWait(self.driver, 100).until(EC.visibility_of_element_located((
                By.XPATH, '//*[@id="mainsearch"]')))
            print('Loading page ...')
            login_button.submit()

        except Exception:
            self.error = True
            print('Could not find search bar')

    def scroll_down(self):
        global position
        try:
            wait = WebDriverWait(self.driver, 10)
            wait.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, '#form-cheapest-acc-hot13445 > button:nth-child(1)')))

        except TimeoutException:
            print("Loading took too much time!")

        try:
            print('Scrolling page ...')
            # self.driver.execute_script('window.scrollTo(0, document.body.scrollHeight);')
            read_mores = self.driver.find_elements_by_xpath('//form/button')
            for read_more in read_mores:
                self.driver.execute_script("arguments[0].scrollIntoView();", read_more)
                # read_more.click()
            sleep(1)

            # even 25 seconds may not be enough to load the page!
            # todo AJAX: maximum 30 seconds
            print('Waiting ajax full load ...')
            sleep(11)  # wait ajax full load

            print("Scraping page ...")
            soup = BeautifulSoup(self.driver.page_source, 'lxml')
            hotel_list = soup.find_all('article', {'class': 'crosselling-line availability-item'})
            euro_symbol = '€'
            print("\n\tSnapshoot:\n")
            try:
                for i, hotel in enumerate(hotel_list):
                    self.all_positions.append(i + 1)

                    hotel_name = hotel.find('a', {'data-tl': 'acc-title'}).getText()
                    # fixme: remove whitespaces REF: https://stackoverrun.com/es/q/743639
                    hotel_name = ' '.join(hotel_name.split())
                    self.all_hotels.append(hotel_name)
                    # print(" %d - %s" % (i + 1, hotel_name))

                    hotel_zone = hotel.find('span', {'class': '_hotelzone'}).getText().strip(',')
                    hotel_name = ' '.join(hotel_name.split())
                    self.all_zones.append(hotel_zone)

                    hotel_integer = hotel.find('span', {'class': 'hotel-price'}).getText()
                    hotel_decimal = hotel.find('span', {'class': 'hotel-price-decimal'}).getText().strip('€')
                    hotel_integer = hotel_integer.replace(',', '')
                    hotel_price = '{}.{}'.format(hotel_integer, hotel_decimal)
                    hotel_price = float(hotel_price)
                    hotel_price = "{0:.2f}".format(hotel_price)
                    self.all_prices.append(hotel_price)
                    if len(hotel_price) == 5:
                        hotel_price = "   " + hotel_price
                    if len(hotel_price) == 6:
                        hotel_price = "  " + hotel_price
                    if len(hotel_price) == 7:
                        hotel_price = " " + hotel_price
                    if len(hotel_price) == 8:
                        hotel_price = "" + hotel_price

                    if i < 9:
                        print(" %d - %s %s %s - %s" % (i + 1, hotel_price, euro_symbol, hotel_name, hotel_zone))
                    else:
                        print("%d - %s %s %s - %s" % (i + 1, hotel_price, euro_symbol, hotel_name, hotel_zone))

                print("\n\tRanking:\n")
                # float cast
                new_prices = []
                for element in self.all_prices:
                    rank = float(element)
                    new_prices.append(rank)

                # final list
                display_list = list(zip(self.all_positions, self.all_hotels, new_prices, self.all_zones))
                ranking = sorted(display_list, key=operator.itemgetter(2))
                # todo REF: https://discuss.codecademy.com/t/how-can-i-sort-a-zipped-object/454412/6
                for j, k, v, w in ranking:
                    if v < 100.00:
                        print("   ", "{0:.2f}".format(v), k)
                    if 99.00 < v < 1000.00:
                        print("  ", "{0:.2f}".format(v), k)
                    if 999.00 < v < 10000.00:
                        print(" ", "{0:.2f}".format(v), k)
                    if v > 9999.00:
                        print("", "{0:.2f}".format(v), k)

                self.cheap = ranking[0]
                self.data = ranking
                print('\nCheapest reservation: ', self.cheap[1], self.cheap[2], euro_symbol)
                # self.display = display_list[7]
                # print('Target button number: ', self.display.index(self.cheap[0]))
                self.display = display_list
                for i, collation in enumerate(display_list):
                    if collation[1] == self.cheap[1]:
                        position = i
                print('Pointing to the target button ', position + 1, ' ...')
                self.index = str(position)
                if self.error is False:
                    self.target_button(self.index)

                sleep(2)
            except Exception as e:
                self.error = True
                print(e)
                print('Some error occurred while trying to scratch the hotel list')

            # activate page analyzer
            if mode == "d":
                print('Downloading html to computer ...')
                file = open("bedsonline.html", "w")
                file.write(soup.prettify())
                file.close()

        except NoSuchElementException:
            print('Some error occurred while trying to scroll down')
            self.error = True

    def target_button(self, index):
        target_button = self.driver.find_element_by_xpath(
            '//article[' + index + ']/div/div[1]/div[4]/div[2]/div[2]/form/button')
        self.driver.execute_script("arguments[0].scrollIntoView();", target_button)
        # target_button.click()

    def file_manager(self, ):
        bookings_folder_path = os.path.join(self.path, 'bookings')
        if not os.path.exists(bookings_folder_path):
            os.mkdir(bookings_folder_path)
        if self.error is False:
            self.write_bookings_to_excel_file(bookings_folder_path)
            print('Writing to excel ...')
        # if self.error is False:
        #     self.read_bookings_from_excel_file(self.path + '/bookings/bookings.xlsx')

    def write_bookings_to_excel_file(self, booking_path):
        # FIXME: openpyxl -> https://openpyxl.readthedocs.io/en/stable/index.html
        filepath = os.path.join(booking_path, 'bookings.xlsx')
        if not os.path.exists(filepath):
            workbook = Workbook()
            workbook.save(filepath)
            # todo: grab the active worksheet: worksheet = workbook.active This is set to 0 by default. Unless you
            #  modify its value, you will always get the first worksheet by using: worksheet = workbook.active -> or
            #  you can create new worksheets using the Workbook.create_sheet() method:
            worksheet_1 = workbook.create_sheet("Snapshoot", 0)  # insert at first position
            worksheet_2 = workbook.create_sheet("Bookings", -1)  # insert at the penultimate position
            # Once you gave a worksheet a name, you can get it as a key of the workbook:
            # >>> ws3 = wb["New Title"]
            # fixme: delete the default sheet:
            std = workbook["Sheet"]
            workbook.remove(std)

            sheet = workbook.active
            self.set_stylesheet(sheet, 1)

            # fixme write REF: https://www.pythonexcel.com/openpyxl-write-to-cell.php
            header = ('No', 'Price', 'Hotel', 'Zone', 'Retail', 'Profit')
            worksheet_1.cell(row=1, column=1).value = header[0]
            worksheet_1.cell(row=1, column=2).value = header[1]
            worksheet_1.cell(row=1, column=3).value = header[2]
            worksheet_1.cell(row=1, column=4).value = header[3]
            worksheet_1.cell(row=1, column=5).value = header[4]
            worksheet_1.cell(row=1, column=6).value = header[5]

            # worksheet_1["E2"] = "=1.374*B2:B31"
            # worksheet_1.formula_attributes['E2'] = {'t': 'array', 'ref': "E2:E31"}
            # fixme: # working with scraped lists you do not know the range length:
            # WARNING: sheet.append(row) only append all rows, use cell instead:
            c = '1.374'
            i = 2
            for row in self.data:
                cell_reference = worksheet_1.cell(row=i, column=1)
                cell_reference.value = row[0]
                cell_reference = worksheet_1.cell(row=i, column=2)
                cell_reference.value = row[2]
                cell_reference = worksheet_1.cell(row=i, column=3)
                cell_reference.value = row[1]
                cell_reference = worksheet_1.cell(row=i, column=4)
                cell_reference.value = row[3]
                # REF:
                # https://stackoverflow.com/questions/51044736/openpyxl-iterate-through-rows-and-apply-formula
                # fixme CODE:
                #  for row_num in range(2, max_row_num):
                #     sheet['E{}'.format(row_num)] = '=CLEAN(D{})'.format(row_num)
                sheet['E{}'.format(i)] = '=PRODUCT(B{},{}'.format(i, c)
                sheet['F{}'.format(i)] = '=SUM(E{},-B{}'.format(i, i)
                i += 1

            # select bookings sheet
            workbook.active = worksheet_2
            sheet = workbook.active
            self.set_stylesheet(sheet, 2)

            header = ('No', 'Price', 'Hotel', 'Zone', 'Retail', 'Profit')
            worksheet_2.cell(row=1, column=1).value = header[0]
            worksheet_2.cell(row=1, column=2).value = header[1]
            worksheet_2.cell(row=1, column=3).value = header[2]
            worksheet_2.cell(row=1, column=4).value = header[3]
            worksheet_2.cell(row=1, column=5).value = header[4]
            worksheet_2.cell(row=1, column=6).value = header[5]

            booking = self.data[0]
            worksheet_2.cell(row=2, column=1).value = booking[0]
            worksheet_2.cell(row=2, column=2).value = booking[2]
            worksheet_2.cell(row=2, column=3).value = booking[1]
            worksheet_2.cell(row=2, column=4).value = booking[3]
            sheet['E{}'.format(2)] = '=PRODUCT(B{},{}'.format(2, c)
            sheet['F{}'.format(2)] = '=SUM(E{},-B{}'.format(2, 2)

            workbook.active = worksheet_1
        else:
            workbook = load_workbook(filepath)
            worksheet_1 = workbook.active

            c = '1.374'
            i = 2
            for row in self.data:
                cell_reference = worksheet_1.cell(row=i, column=1)
                cell_reference.value = row[0]
                cell_reference = worksheet_1.cell(row=i, column=2)
                cell_reference.value = row[2]
                cell_reference = worksheet_1.cell(row=i, column=3)
                cell_reference.value = row[1]
                cell_reference = worksheet_1.cell(row=i, column=4)
                cell_reference.value = row[3]
                # REF:
                # https://stackoverflow.com/questions/51044736/openpyxl-iterate-through-rows-and-apply-formula
                # fixme CODE:
                #  for row_num in range(2, max_row_num):
                #     sheet['E{}'.format(row_num)] = '=CLEAN(D{})'.format(row_num)
                worksheet_1['E{}'.format(i)] = '=PRODUCT(B{},{}'.format(i, c)
                worksheet_1['F{}'.format(i)] = '=SUM(E{},-B{}'.format(i, i)
                i += 1

            workbook.active = 1
            worksheet_2 = workbook.active

            # select target row
            target = 1
            while worksheet_2.cell(row=target, column=1).value is not None:
                target += 1

            booking = self.data[0]
            worksheet_2.cell(row=target, column=1).value = booking[0]
            worksheet_2.cell(row=target, column=2).value = booking[2]
            worksheet_2.cell(row=target, column=3).value = booking[1]
            worksheet_2.cell(row=target, column=4).value = booking[3]
            worksheet_2['E{}'.format(target)] = '=PRODUCT(B{},{}'.format(target, c)
            worksheet_2['F{}'.format(target)] = '=SUM(E{},-B{}'.format(target, target)

            # switch sheet
            workbook.active = 0

        sheet = workbook.active
        self.set_stylesheet(sheet, 1)
        workbook.save(filepath)  # save file

    def set_stylesheet(self, sheet, shift):

        sheet.column_dimensions['B'].number_format = '#,##0.00'
        sheet.column_dimensions['E'].number_format = '#,##0.00'
        sheet.column_dimensions['F'].number_format = '#,##0.00'
        if shift != 2:
            sheet.column_dimensions['A'].width = 3
        else:
            sheet.column_dimensions['A'].width = 4
        sheet.column_dimensions['B'].width = 9
        sheet.column_dimensions['C'].width = 50
        sheet.column_dimensions['D'].width = 16
        sheet.column_dimensions['E'].width = 9
        sheet.column_dimensions['F'].width = 9

        format = sheet.column_dimensions['A']
        format.font = Font(bold=True, italic=True, name='Arial')
        format = sheet.column_dimensions['B']
        format.font = Font(bold=True, italic=True, name='Arial')
        format = sheet.column_dimensions['C']
        format.font = Font(bold=True, italic=True, name='Arial')
        format = sheet.column_dimensions['D']
        format.font = Font(bold=True, italic=True, name='Arial')
        format = sheet.column_dimensions['E']
        format.font = Font(bold=True, italic=True, name='Arial')
        format = sheet.column_dimensions['F']
        format.font = Font(bold=True, italic=True, name='Arial')

        # fixme REF:
        # https://stackoverflow.com/questions/35918504/adding-a-background-color-to-cell-openpyxl
        for col_range in range(1, 7):
            cell_title = sheet.cell(1, col_range)
            cell_title.fill = PatternFill(start_color="00c0c0c0", end_color="00c0c0c0", fill_type="solid")


if __name__ == '__main__':
    mode = input('\nSelect spider run mode:'
                 '\n\tDEMO .........(type "d" + enter)'
                 "\n\tAUTOMATIC ....(press enter)\n")
    if mode != "d":
        x = 0
        while x < 25:
            app = App(target_continent=t.a[x],
                      target_country_col=t.b[x],
                      target_country_row=t.c[x],
                      target_city_col=t.d[x],
                      target_city_row=t.e[x])
            x += 1
    else:
        app = App()
